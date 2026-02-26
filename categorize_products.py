"""
categorize_products.py
======================
Classifies the 2,199 products in "פריטים ללא קבוצת משנה.xlsx" by matching
them to the category taxonomy in "קבוצות קומקס.xlsx".

Output:
  - פריטים_מעודכנים.xlsx  — full updated product list
  - Console summary with stats and suggested new categories

Usage:
  python3 categorize_products.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from collections import defaultdict

TAXONOMY_FILE = "קבוצות קומקס.xlsx"
PRODUCTS_FILE = "פריטים ללא קבוצת משנה.xlsx"
OUTPUT_FILE   = "פריטים_מעודכנים.xlsx"

# ---------------------------------------------------------------------------
# 1. Load taxonomy
# ---------------------------------------------------------------------------

def load_taxonomy(path):
    """Returns a list of dicts, one per row (domain→dept→group→subgroup)."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        domain_id, domain_name, dept_id, dept_name, grp_id, grp_name, sub_id, sub_name, count = row
        if domain_name is None:
            continue
        rows.append({
            "domain_id":   domain_id,
            "domain_name": domain_name,
            "dept_id":     dept_id,
            "dept_name":   dept_name,
            "grp_id":      grp_id,
            "grp_name":    grp_name,
            "sub_id":      sub_id,
            "sub_name":    sub_name,
        })
    return rows


def build_lookup(taxonomy):
    """
    Build lookup structures:
      by_dept_group[(dept, group)] = list of subgroup names
      by_dept[(dept,)] = list of (group, subgroup) pairs
    All keys are lowercased for fuzzy matching.
    """
    by_dept_group = defaultdict(list)
    by_dept       = defaultdict(list)

    for t in taxonomy:
        dept = (t["dept_name"] or "").strip()
        grp  = (t["grp_name"]  or "").strip()
        sub  = (t["sub_name"]  or "").strip()
        if dept and grp and sub:
            by_dept_group[(dept.lower(), grp.lower())].append(t)
        if dept:
            by_dept[dept.lower()].append(t)

    return by_dept_group, by_dept


# ---------------------------------------------------------------------------
# 2. Keyword-based classifier
# ---------------------------------------------------------------------------

# Each entry: (list_of_keywords, (domain, dept, group, subgroup))
# Keywords matched against the product name (Hebrew). Order matters — more
# specific rules FIRST.

RULES = [
    # ── Purim / holidays ──────────────────────────────────────────────────
    (["מגילת אסתר", "מגילות אסתר"],
     ("NF", "חגים", "פורים", "מגילות אסתר")),
    (["תחפושת", "מסכה"],
     ("NF", "חגים", "פורים", "תחפושות ואביזרים")),
    (["משלוח מנות", "שקית פורים", "קופסת פורים"],
     ("NF", "חגים", "פורים", "אביזרים למשלוחי מנות")),
    (["פורים"],           # catch-all after more specific rules
     ("NF", "חגים", "פורים", "אביזרים למשלוחי מנות")),
    (["סביבון", "חנוכיה", "חנוכה"],
     ("NF", "חגים", "חנוכה", "סביבונים")),
    (["שקית חג", "שקיות חג"],
     ("NF", "חגים", "חגים כללי", "שקיות חג")),
    (["מדבקות לפסח", "פסח"],
     ("NF", "חגים", "חגים כללי", "שקיות חג")),   # best fit available

    # ── Religious / Judaica (new category — flagged) ───────────────────────
    (["ציצית", "טלית", "פתיל", "פתילים"],
     ("NF", "יודאיקה *חדש*", "ציצית וטלית *חדש*", "פתילים וציצית *חדש*")),
    (["מזוזה", "תפילין", "שופר"],
     ("NF", "יודאיקה *חדש*", "יודאיקה *חדש*", "יודאיקה *חדש*")),

    # ── Fresh herbs / spices (in the nuts area physically) ─────────────────
    (["חזרת"],
     ("פירות וירקות", "ירקות", "שורשיים", "שורש")),
    (["לוף"],
     ("פירות וירקות", "ירקות", "ירקות לבישול", "ירקות לבישול *חדש*")),
    (["נענע", "כוסברה", "פטרוזיליה", "שמיר", "עשב"],
     ("פירות וירקות", "ירקות", "ירקות עלים", "פטרוזיליה כוסברה נענע שמיר")),

    # ── Nuts & seeds (פיצוחים) — subgroups to CREATE ──────────────────────
    (["בוטן", "בוטנים"],
     ("פירות וירקות", "פיצוחים", "בוטנים *חדש*", "בוטנים *חדש*")),
    (["שקד", "שקדים"],
     ("פירות וירקות", "פיצוחים", "שקדים *חדש*", "שקדים *חדש*")),
    (["פיסטוק"],
     ("פירות וירקות", "פיצוחים", "פיסטוק *חדש*", "פיסטוק *חדש*")),
    (["קשיו"],
     ("פירות וירקות", "פיצוחים", "קשיו *חדש*", "קשיו *חדש*")),
    (["פקאן"],
     ("פירות וירקות", "פיצוחים", "פקאן ומקדמיה *חדש*", "פקאן *חדש*")),
    (["מקדמיה"],
     ("פירות וירקות", "פיצוחים", "פקאן ומקדמיה *חדש*", "מקדמיה *חדש*")),
    (["בונדוק", "אגוז ברזיל", "אגוז"],
     ("פירות וירקות", "פיצוחים", "אגוזים *חדש*", "אגוזים *חדש*")),
    (["גרעיני חמניה", "גרעיני אבטיח", "גרעיני דלעת", "גרעינים"],
     ("פירות וירקות", "פיצוחים", "גרעינים *חדש*", "גרעינים *חדש*")),
    (["גרעינים דלעת", "גרעיני אבטיח זריפה"],
     ("פירות וירקות", "פיצוחים", "גרעינים *חדש*", "גרעינים *חדש*")),
    (["חומוס קלוי"],
     ("פירות וירקות", "פיצוחים", "חומוס ודגנים קלויים *חדש*", "חומוס קלוי *חדש*")),
    (["תירס מטוגן", "תירס"],
     ("פירות וירקות", "פיצוחים", "חטיפי פיצוחים *חדש*", "חטיפי תירס *חדש*")),
    (["קבוקים", "קרנצוס", "רביולי גריל"],
     ("פירות וירקות", "פיצוחים", "חטיפי פיצוחים *חדש*", "חטיפי פיצוחים *חדש*")),
    (["מעורב", "מארז פיצוחים"],
     ("פירות וירקות", "פיצוחים", "תערובות פיצוחים *חדש*", "תערובות פיצוחים *חדש*")),
    (["ממתק פרי"],
     ("פירות וירקות", "פיצוחים", "ממתקי פיצוחים *חדש*", "ממתקי פיצוחים *חדש*")),

    # ── Fresh fish ─────────────────────────────────────────────────────────
    (["דג טרי", "פילה", "דניס", "לברק", "אמנון", "קרפיון", "סלמון", "טונה טרי",
      "אינטיאס", "ברי", "פרידה", "כוקיה", "מוסר", "בקלה", "פגריה"],
     ("מצוננים", "דגים טריים", "דגים טריים", "דגים שלמים טריים *חדש*")),
    (["דג ארוז", "פילה ארוז", "דג מנוקה"],
     ("מצוננים", "דגים טריים", "דגים טריים ארוזים", "דגים ארוזים *חדש*")),

    # ── Meat ───────────────────────────────────────────────────────────────
    (["אנטריקוט", "סינטה", "פילה בקר", "אסאדו", "שפונדרה", "צלעות בקר",
      "כתף בקר", "שייטל", "לשון"],
     ("מצוננים", "קצביה בקר טרי", "חלקי בשר טרי", "נתחי בקר *חדש*")),
    (["בשר טחון", "קציצות"],
     ("מצוננים", "קצביה בקר טרי", "בשר טרי טחון", "בשר טחון *חדש*")),
    (["בשר ארוז"],
     ("מצוננים", "קצביה בקר טרי", "בשר טרי ארוז", "בשר ארוז *חדש*")),
    (["עוף שלם", "פרגית שלמה"],
     ("מצוננים", "קצביה עופות טריים", "עוף טרי שלם", "עוף שלם *חדש*")),
    (["חזה עוף", "שוק עוף", "כנף עוף", "ירך עוף", "חלקי עוף"],
     ("מצוננים", "קצביה עופות טריים", "חלקי עוף טרי", "חלקי עוף *חדש*")),
    (["הודו", "פרגית הודו", "חזה הודו", "שוק הודו"],
     ("מצוננים", "קצביה עופות טריים", "הודו טרי", "חלקי הודו *חדש*")),
    (["עוף טחון", "הודו טחון"],
     ("מצוננים", "קצביה עופות טריים", "עוף והודו טחון", "עוף טחון *חדש*")),

    # ── Textiles ───────────────────────────────────────────────────────────
    (["מצעים", "סדין", "ציפית"],
     ("NF", "טקסטיל", "כלי מיטה", "מצעים")),
    (["שמיכה", "שמיכות"],
     ("NF", "טקסטיל", "כלי מיטה", "שמיכות קיץ")),
    (["כרית", "כריות"],
     ("NF", "טקסטיל", "כלי מיטה", "כריות")),
    (["מגבת", "מגבות"],
     ("NF", "טקסטיל", "מגבות", "מגבות גוף")),
    (["מפה", "מפות"],
     ("NF", "טקסטיל", "טקסטיל לבית", "מפות")),
    (["שטיח"],
     ("NF", "טקסטיל", "טקסטיל לבית", "שטיחים")),
    (["גרב", "גרביים"],
     ("NF", "טקסטיל", "ביגוד", "גרבי גברים")),
    (["צעיף"],
     ("NF", "טקסטיל", "אביזרי חורף", "צעיפים")),
    (["כפפות"],
     ("NF", "טקסטיל", "אביזרי חורף", "כפפות")),
    (["מטריה"],
     ("NF", "טקסטיל", "אביזרי חורף", "מטריות")),

    # ── Household / NF ─────────────────────────────────────────────────────
    (["כבל", "תקע", "שקע"],
     ("NF", "מוצרי חשמל", "אביזרי חשמל ותאורה", "תקעים וכבלים")),
    (["נורה", "גוף תאורה"],
     ("NF", "מוצרי חשמל", "אביזרי חשמל ותאורה", "נורות וגופי תאורה")),
    (["סוללה", "סוללות"],
     ("NF", "מוצרי חשמל", "אלקטרוניקה", "סוללות")),
    (["אוזניות"],
     ("NF", "מוצרי חשמל", "אלקטרוניקה", "אוזניות")),
    (["מטען", "כבל טעינה"],
     ("NF", "מוצרי חשמל", "אלקטרוניקה", "מטענים וכבלים")),
    (["מברשת שיניים", "חוט דנטלי", "מגרד לשון", "קיסם"],
     ("פארם", "היגיינת הפה", "מברשות שיניים", "מברשות שיניים *חדש*")),
    (["פח", "פחים"],
     ("NF", "כלי בית", "מוצרים לבית", "פחים")),
    (["מעצור דלת"],
     ("NF", "כלי בית", "מוצרים לבית", "מוצרים לבית *חדש*")),
    (["פרלטור", "ברז", "צינור"],
     ("NF", "כלי בית", "מוצרים לאמבטיה", "ברזים וחסכמים")),
    (["בלון"],
     ("NF", "פנאי", "אביזרי מסיבה", "בלונים")),
    (["פנקס", "מחברת"],
     ("NF", "פנאי", "ציוד משרדי", "מחברות בלוקים ומעטפות")),
    (["מדבקות"],
     ("NF", "פנאי", "יצירה", "מדבקות")),
    (["צנצנת", "קוצץ ציפורניים"],
     ("NF", "כלי בית", "כלי אחסון", "מוצרי אחסון פלסטיק")),

    # ── Dry food / spices ──────────────────────────────────────────────────
    (["בצל מטוגן", "בצל פריך"],
     ("מזון יבש", "מוצרים לבישול ואפיה", "תבלינים", "תבלינים בשקית")),

    # ── Beverages ──────────────────────────────────────────────────────────
    (["מיץ", "נקטר"],
     ("משקאות", "משקאות קלים", "נקטרים ומיצים", "מיצים *חדש*")),

    # ── Garden / plants ────────────────────────────────────────────────────
    (["לוונדר", "צמח", "עציץ", "פרח"],
     ("NF", "גינה *חדש*", "צמחים *חדש*", "צמחים וזרעים *חדש*")),
]


def match_by_keyword(name: str):
    """Returns (domain, dept, group, subgroup) or None."""
    name_lower = name.lower() if name else ""
    for keywords, category in RULES:
        for kw in keywords:
            if kw.lower() in name_lower:
                return category
    return None


# ---------------------------------------------------------------------------
# 3. Subgroup finder from taxonomy
# ---------------------------------------------------------------------------

def find_subgroup(dept_name, group_name, product_name, taxonomy):
    """
    Given existing dept + group, find the best matching subgroup from taxonomy.
    Returns subgroup name string or None.
    """
    matches = [t for t in taxonomy
               if (t["dept_name"] or "").strip() == dept_name
               and (t["grp_name"]  or "").strip() == group_name
               and t["sub_name"]]

    if not matches:
        return None

    # Only one option → use it
    if len(matches) == 1:
        return matches[0]["sub_name"]

    # Multiple options → try keyword matching on product name
    name_lower = (product_name or "").lower()
    for m in matches:
        sub_lower = (m["sub_name"] or "").lower()
        # Check if any word in the subgroup name appears in the product name
        words = sub_lower.split()
        if any(w in name_lower for w in words if len(w) > 2):
            return m["sub_name"]

    # Fallback: return first match
    return matches[0]["sub_name"]


# ---------------------------------------------------------------------------
# 4. Main classification
# ---------------------------------------------------------------------------

def classify_products(taxonomy, products_ws):
    results = []
    stats = {"auto_full": 0, "auto_subgroup": 0, "new_category": 0,
             "skipped": 0, "unknown": 0}
    new_categories = set()

    for row in products_ws.iter_rows(min_row=2, values_only=True):
        item_id, item_name, domain, dept, group, subgroup, supplier_id, supplier_name = row

        # Skip placeholder items
        if item_name and "פריט חדש" in str(item_name):
            stats["skipped"] += 1
            results.append({
                "item_id": item_id, "item_name": item_name,
                "domain": domain, "dept": dept, "group": group,
                "subgroup": "פריט לא קיים - לדילוג",
                "supplier_id": supplier_id, "supplier_name": supplier_name,
                "status": "skipped"
            })
            continue

        # Case 1: Has domain + dept + group → find subgroup from taxonomy
        if dept and group:
            found_sub = find_subgroup(dept, group, item_name, taxonomy)
            if found_sub:
                stats["auto_subgroup"] += 1
                results.append({
                    "item_id": item_id, "item_name": item_name,
                    "domain": domain, "dept": dept, "group": group,
                    "subgroup": found_sub,
                    "supplier_id": supplier_id, "supplier_name": supplier_name,
                    "status": "auto_subgroup"
                })
                continue

        # Case 2: Has dept but no group → try keyword match
        # Case 3: No dept at all → try keyword match
        kw_result = match_by_keyword(item_name)
        if kw_result:
            new_domain, new_dept, new_group, new_sub = kw_result
            is_new = "*חדש*" in " ".join([new_dept, new_group, new_sub])
            if is_new:
                new_categories.add((new_domain, new_dept, new_group, new_sub))
                stats["new_category"] += 1
                status = "new_category"
            else:
                stats["auto_full"] += 1
                status = "auto_full"

            results.append({
                "item_id": item_id, "item_name": item_name,
                "domain": new_domain, "dept": new_dept,
                "group": new_group, "subgroup": new_sub,
                "supplier_id": supplier_id, "supplier_name": supplier_name,
                "status": status
            })
            continue

        # Case 4: Has dept but product-level subgroup lookup also failed
        if dept and not group:
            # Try to find any subgroup under the dept from taxonomy
            dept_matches = [t for t in taxonomy
                            if (t["dept_name"] or "").strip() == dept and t["sub_name"]]
            if dept_matches:
                t = dept_matches[0]
                stats["auto_subgroup"] += 1
                results.append({
                    "item_id": item_id, "item_name": item_name,
                    "domain": domain or t["domain_name"],
                    "dept": dept,
                    "group": t["grp_name"],
                    "subgroup": t["sub_name"] + " *לבדיקה*",
                    "supplier_id": supplier_id, "supplier_name": supplier_name,
                    "status": "partial_auto"
                })
                continue

        # Unknown — needs manual review
        stats["unknown"] += 1
        results.append({
            "item_id": item_id, "item_name": item_name,
            "domain": domain or "לא ידוע",
            "dept": dept or "לבדיקה",
            "group": group or "לבדיקה",
            "subgroup": "לבדיקה ידנית",
            "supplier_id": supplier_id, "supplier_name": supplier_name,
            "status": "unknown"
        })

    return results, stats, new_categories


# ---------------------------------------------------------------------------
# 5. Write output Excel
# ---------------------------------------------------------------------------

STATUS_COLORS = {
    "auto_full":      "C6EFCE",  # green
    "auto_subgroup":  "DDEBF7",  # light blue
    "new_category":   "FFEB9C",  # yellow — new category needed
    "partial_auto":   "FCE4D6",  # orange — review recommended
    "unknown":        "FFCCCC",  # red — manual review required
    "skipped":        "D9D9D9",  # grey
}

STATUS_LABELS = {
    "auto_full":     "סווג אוטומטי",
    "auto_subgroup": "קבוצת משנה אוטומטית",
    "new_category":  "קטגוריה חדשה נדרשת",
    "partial_auto":  "לבדיקה - חלקי",
    "unknown":       "לבדיקה ידנית",
    "skipped":       "פריט לא קיים",
}


def write_output(results, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "פריטים מעודכנים"

    headers = [
        "פריט", "שם פריט", "תאור מחלקת על", "שם מחלקה",
        "שם קבוצה", "שם קבוצת משנה", "ספק", "שם ספק", "סטטוס"
    ]
    ws.append(headers)

    # Header styling
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in results:
        row_data = [
            r["item_id"], r["item_name"],
            r["domain"], r["dept"], r["group"], r["subgroup"],
            r["supplier_id"], r["supplier_name"],
            STATUS_LABELS.get(r["status"], r["status"])
        ]
        ws.append(row_data)

        # Color coding by status
        color = STATUS_COLORS.get(r["status"], "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        row_idx = ws.max_row
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
    print(f"\nOutput saved: {output_path}")


# ---------------------------------------------------------------------------
# 6. Main
# ---------------------------------------------------------------------------

def main():
    print("Loading taxonomy...")
    taxonomy = load_taxonomy(TAXONOMY_FILE)
    print(f"  Loaded {len(taxonomy)} taxonomy entries.")

    print("Loading products...")
    wb2 = openpyxl.load_workbook(PRODUCTS_FILE)
    ws2 = wb2.active
    print(f"  Loaded {ws2.max_row - 1} products.")

    print("Classifying products...")
    results, stats, new_categories = classify_products(taxonomy, ws2)

    print("\n── Classification Summary ──────────────────────────────────")
    total = sum(stats.values())
    for k, v in stats.items():
        pct = v / total * 100 if total else 0
        print(f"  {STATUS_LABELS.get(k, k):<30} {v:>5}  ({pct:.1f}%)")
    total_label = 'סה"כ'
    print(f"  {total_label:<30} {total:>5}")

    if new_categories:
        print("\n── New Categories to Create in System ──────────────────────")
        print("  (marked with *חדש* in the output file)")
        for nc in sorted(new_categories):
            domain, dept, group, sub = nc
            print(f"  תחום:{domain} | מחלקה:{dept} | קבוצה:{group} | ק.משנה:{sub}")

    write_output(results, OUTPUT_FILE)

    print("\n── Color Legend ────────────────────────────────────────────")
    print("  🟢 ירוק    — סווג אוטומטי מלא (בטוח)")
    print("  🔵 כחול    — קבוצת משנה הוספה אוטומטית")
    print("  🟡 צהוב    — קטגוריה חדשה נדרשת (צריך ליצור במערכת)")
    print("  🟠 כתום    — סווג חלקי — לבדיקה")
    print("  🔴 אדום    — לבדיקה ידנית (לא זוהה)")
    print("  ⬜ אפור    — פריט לא קיים / לדילוג")


if __name__ == "__main__":
    main()
